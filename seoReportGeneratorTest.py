from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from webdriver_utility import WebDriverUtility
from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
import os 
import re
import time 
from datetime import datetime 
from urllib.parse import urlparse


class SEOReportGenerator: 

	def __init__(self):
		# Define the base path where Excel files will be saved
		self.base_path = "/Users/jonathanbachrach/Projects/SEOReportGenerator"

		# Initialize the WebDriver using WebDriverUtility
		self.driver = WebDriverUtility.setup_driver()
		if not self.driver:
			print("Failed to initialize the WebDriver. Please check the setup.")
			exit(1)  # stop program execution if the driver fails to initialize
			# Define the base path where Excel files will be saved

		# Pass the WebDriver instance to BrowserNavigator
		self.browser_navigator = BrowserNavigator(self.driver, self.base_path)

		self.data_parser = DataParser()
		self.excel_manager = ExcelManager(self.base_path)  # Pass base path to ExcelManager


	def login(self):
		"""Log into Google Search Console with the provided username and password."""
		print("Navigating to the login page...")
		login_url = 'https://accounts.google.com/v3/signin/identifier?continue=https%3A%2F%2Fsearch.google.com%2Fu%2F2%2Fsearch-console%2Findex%3Fresource_id%3Dsc-domain%3Aturnkeyofficespace.com&followup=https%3A%2F%2Fsearch.google.com%2Fu%2F2%2Fsearch-console%2Findex%3Fresource_id%3Dsc-domain%3Aturnkeyofficespace.com&ifkv=ARZ0qKJdYymqdmqQWq01fihTEzH96aEB97nVtP7y5_8bAPXyPMZ2KrHmWXy1eEX9qZ6RKuPmPG7iCg&passive=1209600&service=sitemaps&flowName=GlifWebSignIn&flowEntry=ServiceLogin&dsh=S-620576232%3A1712399017622814&theme=mn&ddm=0'
		self.driver.get(login_url)
		print("Page loaded, attempting to fill username...")

		# Input username
		user_elem = WebDriverWait(self.driver, 10).until(
			EC.presence_of_element_located((By.ID, "identifierId"))
		)

		user_elem.send_keys('jon@turnkeyofficespace.com')
		
		# Click the 'Next" button after entering the username
		next_button = WebDriverWait(self.driver, 10).until(
			EC.element_to_be_clickable((By.CSS_SELECTOR, "#identifierNext > div > button > span"))
		)
		next_button.click()
		print("Username entered, waiting for password filed...")
		time.sleep(15)

		# Wait for transition and input password 
		password_elem = WebDriverWait(self.driver, 10).until(
			EC.presence_of_element_located((By.NAME, "Passwd"))
		)
		time.sleep(2)
		print("Password field available, entering password...")

		password_elem.send_keys('Kjwmx6Koaet2jx')

		# Click the 'Next" button after entering the password
		next_button_password = WebDriverWait(self.driver, 10).until(
			EC.element_to_be_clickable((By.CSS_SELECTOR, "#passwordNext > div > button > span"))
		)

		next_button_password.click()
		print("Password submitted, waiting for 2FA prompt...")

		# Manually handle 2FA
		self.enter_two_factor_code()

		time.sleep(2)

	def enter_two_factor_code(self): 
		try: 
			# Wait for the 2FA code input field to appear 
			code_input = WebDriverWait(self.driver, 30).until(
				EC.presence_of_element_located((By.CSS_SELECTOR, 
					"input[type='text'][autocomplete='one-time-code']"))
			)
			verification_code = input("Enter your 2FA code: ")
			code_input.send_keys(verification_code)

			# Locate and click Next button after entering 2FA code
			next_button = WebDriverWait(self.driver, 10).until(
				EC.element_to_be_clickable((By.CSS_SELECTOR, 
					"#idvPreregisteredPhoneNext > div > button > span"))
			)
			next_button.click()

		except TimeoutException:
			print("2FA input field not found or Next button not clickable.")
			self.driver.save_screenshot('2fa_error.png')  # save a screenshot for debugging 

	# At what point in the following workflow should "Monthly_SEO_Metrics.xlsx" be created?
	def run(self): 
		# Initialize or open the main workbook
		monthly_metrics_path = os.path.join(self.base_path, "Monthly_SEO_Metrics.xlsx")
		workbook = load_workbook(monthly_metrics_path) if os.path.exists(monthly_metrics_path) else Workbook()
		workbook.save(monthly_metrics_path)  # Save immediately to ensure it exists for further processing

		# Navigate to the initial URL of Google Search Console
		self.browser_navigator.navigate_to_console()

		"""HTML content is obtained and passed to DataParser for each type of data required"""

		# Indexed Pages Data
		indexed_pages = self.browser_navigator.get_indexed_pages()
		indexed_data = self.data_parser.parse_indexed_pages(indexed_pages)
		self.excel_manager.update_indexed_pages(indexed_data)
		self.excel_manager.copy_indexed_pages()
		# self.excel_manager._copy_data('Indexed_Pages.xlsx', 'Monthly_SEO_Metrics.xlsx', 'Indexed Pages')

		# 404s Data 
		all_404_urls = self.browser_navigator.get_404_urls()
		valid_404_urls = self.data_parser.parse_404_urls(all_404_urls)
		self.excel_manager.write_404_urls(valid_404_urls)

		# Performance Data (Queries and Top Pages)
		latest_file = self.browser_navigator.get_performance_data()
		self.excel_manager.write_performance_data(latest_file)

		# Total Clicks Last 3 Months Data
		total_organic_clicks = self.browser_navigator.get_total_clicks()
		total_clicks_data = self.data_parser.parse_total_clicks_data(total_organic_clicks)
		self.excel_manager.update_total_clicks_data(total_clicks_data)
		self.excel_manager.copy_total_clicks()
		# self.excel_manager.copy_total_clicks()  # copy updated Total Clicks to Monthly_SEO_Metrics.xlsx


		# Save the workbook after all updates
		# workbook.save(monthly_metrics_path)


	def close(self): 
		"""Method to close the WebDriver when done"""
		self.browser_navigator.driver.quit()


class BrowserNavigator:


	def __init__(self, driver, download_path):
		self.driver = driver 
		self.download_path = "/Users/jonathanbachrach/Downloads"

	def navigate_to_console(self): 
		current_url = self.driver.current_url
		target_url = 'https://search.google.com/u/2/search-console/index?resource_id=sc-domain:turnkeyofficespace.com'
		if current_url != target_url: 
			self.driver.get(target_url)
		try: 
			# Wait for a specific element that signifies the page has loaded
			WebDriverWait(self.driver, 10).until(
				EC.presence_of_element_located((By.CSS_SELECTOR, ".nnLLaf.vtZz6e"))
			)
			print("Navigatied to console and page in loaded.")
		except TimeoutException:
			print("Failed to load the Google Search Console dashboard properly.")
			self.driver.save_screenshot('console_load_fail.png')


	def get_indexed_pages(self): 
		"""Returns indexed_data dict{}"""
		# Navigate to the URL that contains the indexed pages info 
		self.navigate_to_console() 

		indexed_data = {"Last Updated": None, "Indexed Count": None}

		# Retrieve all elements matching the CSS selector
		elements = self.driver.find_elements(By.CSS_SELECTOR, ".nnLLaf.vtZz6e")

		if len(elements) >= 2:  # Ensure there are at least two elements
			indexed_element = elements[1]  # Assuming Indexed Count comes after Not indexed
			indexed_data["Indexed Count"] = indexed_element.get_attribute('title')

			try: 
				# Locate the element that includes "Last Updated" text using the Indexed element
				last_updated_element = self.driver.find_element(By.XPATH, "//*[contains(text(), 'Last updated:')]")
				full_text = last_updated_element.find_element(By.XPATH, "./..").text

				# Extract the date using regex
				match = re.search(r'\d{1,2}/\d{1,2}/\d{2}', full_text)
				if match: 
					indexed_data["Last Updated"] = match.group(0)

			except NoSuchElementException:
				print("Last updated element not found.")
		else: 
			print("Failed to find enough data elements for Indexed Count and Last Updated.")

		print(f"Found data: {indexed_data}")

		return indexed_data 


	def wait_for_download_complete(self, filename_prefix, timeout=25):
		start_time = time.time()
		initial_delay = 10  # Wait 10 seconds before the first check
		time.sleep(initial_delay)  # Allow some time for the download to initiate
		while True:
			files = [f for f in os.listdir(self.download_path) if f.startswith(filename_prefix) and f.endswith('.xlsx')]
			if files: 
				return os.path.join(self.download_path, files[0])
			elif time.time() - start_time > timeout:
				self.driver.save_screenshot('timeout_failure.png')
				raise TimeoutException("Timed out waiting for download to complete.")
			time.sleep(1)


	def get_404_urls(self): 
		# Navigate to page listing 404s
		url = 'https://search.google.com/u/2/search-console/index/drilldown?resource_id=sc-domain%3Aturnkeyofficespace.com&item_key=CAMYDSAC'
		self.driver.get(url)
		WebDriverWait(self.driver, 20).until(
			EC.presence_of_element_located((By.CSS_SELECTOR, ".izuYW"))
		)

		# Click the EXPORT button
		export_button = self.driver.find_element(By.CSS_SELECTOR, "span.izuYW")
		export_button.click()

		# Wait and click 'Download Excel'
		WebDriverWait(self.driver, 10).until(
			EC.visibility_of_element_located((By.XPATH, "//div[text()='Download Excel']"))
		).click()

		# Use the wait_for_download_complete method to ensure the file is fully downloaded
		filename_prefix = "turnkeyofficespace.com-Coverage-Drilldown"
		latest_file = self.wait_for_download_complete(filename_prefix)

		# Open and read the Excel file
		wb = load_workbook(latest_file)
		sheet = wb["Table"]
		all_404_urls = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}


		# Optionally, remove the file if no longer needed
		os.remove(latest_file)

		return all_404_urls


	def get_performance_data(self): 
		"""Method responsible for opening downloaded workbook, extracting data 
		from 'Queries' and 'Pages' sheets, and writing them to the 
		'Monthly_SEO_Metrics.xlsx' workbook."""
		try: 
			url = 'https://search.google.com/u/0/search-console/performance/search-analytics?resource_id=sc-domain:turnkeyofficespace.com&breakdown'
			self.driver.get(url)

			print("Navigating to the performance page...")
			WebDriverWait(self.driver, 20).until(
				EC.presence_of_element_located((By.CSS_SELECTOR, ".izuYW"))
			)
			print("Attempting to find and click the EXPORT button...")
			export_button = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "span.izuYW")))
			export_button.click()
			print("EXPORT button clicked.")

			print("Attempting to find and click 'Download Excel'...")
			download_excel_button = WebDriverWait(self.driver, 10).until(
				EC.element_to_be_clickable((By.XPATH, "//div[text()='Download Excel']"))
			)
			download_excel_button.click()
			print("Download Excel clicked.")

			filename_prefix = "turnkeyofficespace.com-Performance-on-Search"
			latest_file = self.wait_for_download_complete(filename_prefix)

			if latest_file:
				print(f"File downloaded: {latest_file}")
				return latest_file
			else: 
				raise Exception("Download failed, no file found.")

		except Exception as e: 
			print(f"Error during download initiation: {str(e)}")
			return None


	def get_total_clicks(self): 
		"""Returns the dictionary total_clicks_data with the current date as the key."""
		self.driver.get(
			'https://search.google.com/u/2/search-console/performance/search-analytics'
			'?resource_id=sc-domain%3Aturnkeyofficespace.com')

		# Assure chart and page have loaded
		WebDriverWait(self.driver, 10).until(
			EC.presence_of_element_located((By.CSS_SELECTOR, ".nnLLaf.vtZz6e"))
		)

		total_clicks_data = {}

		# Find the element containing the total clicks using its unique class names
		try:
			total_clicks_element = self.driver.find_element(By.CSS_SELECTOR, "div.nnLLaf.vtZz6e")
			total_clicks_title = total_clicks_element.get_attribute('title')
			# The title attribute contains the numerical value as a string
			# Convert the title string to an integer
			total_clicks = int(total_clicks_title.replace(',', ''))  # Remove commas from thousands

			# Get today's date in mm/dd/yy
			current_date = datetime.now().strftime("%m/%d/%y")
			total_clicks_data = {"Last Updated": current_date, "Total Clicks": total_clicks}

			print(f"Total clicks as of {current_date}: {total_clicks}")
		
		except NoSuchElementException:
			print("Total clicks element not found.")
		except ValueError: 
			print("Error processing total clicks data.")

		return total_clicks_data

"""
@staticmethod decorator allows you to call the directly using the class name, without creating an instance.
For example class.method(parameter1, parameter2)
Static methods are often used for utility or helper tasks that don't need to access or 
modify the state of a class instance.
"""

class DataParser:

	@staticmethod
	def is_valid_url(url): 
		try:
			result = urlparse(url)
			# Check if the url has a scheme (eg http or https) and a netloc (domain name)
			return all([result.scheme, result.netloc])
		except Exception: 
			return False

		"""
		Example usage: 
		print(is_valid_url("https://www.example.com"))  # Output: True
		print(is_valid_url("www.example.com"))  # Output: False
		print(is_valid_url("example"))  # Output: False
		"""

	@staticmethod
	def validate_urls(urls):
		# Use is_valid_url to filter and return only valid URLs
		return [url for url in urls if is_valid_url(url)]


	@staticmethod
	def parse_indexed_pages(indexed_data):
		"""Basic validation/transformation for indexed pages data.
		'data' might be a dict with 'Indexed Count' and 'Last Updated'
		"""
		if "Indexed Count" in indexed_data and isinstance(indexed_data["Indexed Count"], str):
			indexed_data["Indexed Count"] = int(indexed_data["Indexed Count"].replace(',', ''))
		if "Last Updated" in indexed_data and isinstance(indexed_data["Indexed Count"], str):
			# Placeholder for date parsing if necessary
			pass
		print("Parsed Indexed Pages Data:", indexed_data)

		return indexed_data


	@staticmethod
	def parse_404_urls(all_404_urls):
		"""Simply returns the dictionary of 404 URLs and their last crawled dates"""
		if not all_404_urls:
			print("No URLs provided to parse.")
			return {}

		valid_404_urls = {} 
		for url, last_crawled in all_404_urls.items(): 
			if DataParser.is_valid_url(url):
				valid_404_urls[url] = last_crawled

		print("Valid 404 URLs:", valid_404_urls)
		return valid_404_urls

	
	@staticmethod
	def parse_total_clicks_data(total_clicks_data):
		# Additional checks performed here if needed
		print("Parsed Total Clicks Data:", total_clicks_data)
		return total_clicks_data


class ExcelManager: 

	def __init__(self, base_path): 
		self.base_path = base_path  # Directory to save Excel files
		self.workbooks = {}  # Dictionary to hold open workbooks


	def update_indexed_pages(self, indexed_data): 
		"""Updates or creates 'Indexed_Pages.xlsx' with indexed page data."""

		filepath = os.path.join(self.base_path, 'Indexed_Pages.xlsx')
		if os.path.exists(filepath):
			wb = load_workbook(filepath)
			# Check if the 'Indexed Data' sheet exists, use it, otherwise create it
			if 'Indexed Data' in wb.sheetnames:
				ws = wb['Indexed Data']
			else:
				ws = wb.create_sheet('Indexed Data')
				ws.append(["Last Updated", "Indexed Count"])  # set desired headers
		else: 
			wb = Workbook()
			ws = wb.create_sheet("Indexed Data")
			ws.append(["Last Updated", "Indexed Count"])			

		# Adding new data under the headers
		ws.append([indexed_data["Last Updated"], indexed_data["Indexed Count"]])
		wb.save(filepath)
		
	def update_total_clicks_data(self, total_clicks_data):
		"""Updates or creates an Excel workbook for total clicks data within the provided workbook."""
		filepath = os.path.join(self.base_path, 'Total_Clicks.xlsx')
		# Load the workbook if it exists, otherwise create a new one
		if os.path.exists(filepath):
			wb = load_workbook(filepath)
		else:
			wb = Workbook()
			ws = wb.create_sheet("Total Clicks")
			ws.append(["Last Updated", "Total Clicks"])  # Ensure headers match dictionary keys

		# Check if the 'Total Clicks' sheet exists, otherwise create it
		if 'Total Clicks' in wb.sheetnames:
			ws = wb['Total Clicks']
		else:
			ws = wb.create_sheet('Total Clicks')
			ws.append(["Last Updated", "Total Clicks"])  # Set headers

		# Append new data under the headers
		ws.append([total_clicks_data["Last Updated"], total_clicks_data["Total Clicks"]])
		wb.save(filepath)  # Save the workbook after modifications
 

	def write_404_urls(self, valid_404_urls):
		"""Writes validated 404 urls into 'Monthly_SEO_Metrics.xlsx'."""
		filepath = os.path.join(self.base_path, 'Monthly_SEO_Metrics.xlsx')
		wb = load_workbook(filepath) if os.path.exists(filepath) else Workbook()
		ws = wb.get_sheet_by_name('404s') if '404s' in wb.sheetnames else wb.create_sheet('404s')
		ws.append(['URL', 'Last Crawled'])
		for url, last_crawled in valid_404_urls.items():
			ws.append([url, last_crawled])
		wb.save(filepath)

		
	def write_performance_data(self, latest_file):
		"""Write performance data from downloaded file into the specified workbook."""
		# Load the performance_data workbook
		try:
			performance_wb = load_workbook(latest_file)
			queries_sheet = performance_wb['Queries']
			pages_sheet = performance_wb['Pages']

			# Headers for each sheet
			queries_headers = ["Top queries", "Clicks", "Impressions", "CTR", "Position"]
			pages_headers = ["Top pages", "Clicks", "Impressions", "CTR", "Position"]

			# Write data to 'Queries Last 3 Months' and 'Top Pages Last 3 Months'
			self._write_sheet_data("Monthly_SEO_Metrics.xlsx", "Queries Last 3 Months", 
				queries_sheet, queries_headers)

			self._write_sheet_data("Monthly_SEO_Metrics.xlsx", "Top Pages Last 3 Months", 
				pages_sheet, pages_headers)

			# Remove the downloaded file
			os.remove(latest_file)
			print("Downloaded file removed and performance data successfully written.")

		except Exception as e:
			print(f"Error processing performance data: {e}")
			# Optionally remove the file even on failure if no further analysis or retries are intended
			if os.path.exists(latest_file):
				os.remove(latest_file)
				print("Downloaded file removed after encountering an error.")


	def _write_sheet_data(self, workbook_name, sheet_title, source_sheet, headers):
		# Load or create the workbook 
		try:
			# Load the destination workbook or create a new one if it doesn't exist
			wb = load_workbook(workbook_name)
		except FileNotFoundError:
			wb = Workbook()
			wb.remove(wb.active)  # Remove default sheet if new workbook

		# Get or create the sheet within the workbook
		if sheet_title in wb.sheetnames:
			ws = wb[sheet_title]
		else:
			ws = wb.create_sheet(title=sheet_title)
			# Set headers if new sheet
			ws.append(headers)

		# Write rows from the source sheet, starting from second row
		for row in source_sheet.iter_rows(min_row=2, values_only=True):  # skips header row
			ws.append(row)

		# Save the changes 
		wb.save(workbook_name)
		print(f"Data written to sheet {sheet_title} in {workbook_name}")


	def copy_indexed_pages(self):
		"""Copies indexed pages data from the Indexed_Pages.xlsx to the 
		Monthly_SEO_Metrics.xlsx workbook."""
		self._copy_data("Indexed_Pages.xlsx", "Monthly_SEO_Metrics.xlsx", "Indexed Data", "Indexed Pages")


	def copy_total_clicks(self): 
		"""Copies total clicks data from the Total_Clicks.xlsx to the 
		Monthly_SEO_Metrics.xlsx workbook."""
		self._copy_data("Total_Clicks.xlsx", "Monthly_SEO_Metrics.xlsx", "Total Clicks", "Total Clicks Last 3 Months")


	def save_workbook(self, workbook_name): 
		"""Saves the workbook after all updates or changes have been made."""
		filepath = os.path.join(self.base_path, workbook_name)
		wb = load_workbook(filepath)
		
		# Check if the default sheet "Sheet" is empty and remove it
		if "Sheet" in wb.sheetnames and all(cell.value is None for row in wb['Sheet'] for 
			cell in row):
			std_wb = wb['Sheet']
			wb.remove(std_wb)

		wb.save(filepath)
		print(f"{workbook_name} saved and cleaned up.")


	def _update_workbook(self, filepath, data, headers):
		"""Helper method to update or create a new Excel workbook with data."""
		# Check if the file exists, if not, create a new workbook and sheet with headers
		if os.path.exists(filepath): 
			wb = load_workbook(filepath)
			ws = wb.active 
		else: 
			wb = Workbook()
			ws = wb.active 
			ws.append(headers)  # add headers if new workbook

		# Append new data
		for key, value in data.items():
			ws.append([key, value])
		wb.save(filepath)


	def _copy_data(self, src_file, dest_file, src_sheet_name, dest_sheet_name):
		"""Copies data from source file to destination file, to appropriate sheet."""
		src_path = os.path.join(self.base_path, src_file)
		dest_path = os.path.join(self.base_path, dest_file)

		print(f"Loading source workbook from {src_path}")
		src_wb = load_workbook(src_path)
		src_ws = src_wb[src_sheet_name]

		print(f"Loading destination workbook from {dest_path}")
		dest_wb = load_workbook(dest_path) if os.path.exists(dest_path) else Workbook()

		if dest_sheet_name in dest_wb.sheetnames:
			dest_ws = dest_wb[dest_sheet_name]
		else:
			dest_ws = dest_wb.create_sheet(title=dest_sheet_name)
			# Assuming the first row in the source sheet contains headers
			headers = [cell.value for cell in src_ws[1]]
			dest_ws.append(headers)

		# Copy data from source to destination starting from the second row to avoid headers
		for row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row, values_only=True):
			dest_ws.append(row)

		print(f"Saving data to {dest_sheet_name} in {dest_file}")
		dest_wb.save(dest_path)
		print(f"Data copied sucessfully from {src_sheet_name} in {src_file} to {dest_sheet_name} in {dest_file}.")
		src_wb.close()  # Ensure the workbook is closed after operation



report_generator = SEOReportGenerator()
report_generator.login()
report_generator.run()
report_generator.close()

	

